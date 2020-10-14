from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet('aaa')
ws = wb.create_sheet('aaass')


def add_data(df, symbol, ws):
    ws.append([])
    ws.append([symbol])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


data = [['Alessx', 10], ['Bob', 12], ['Clarke', 13]]
df = pd.DataFrame(data, columns=['Name', 'Age'])


def aa(p):
    """
    docstring
    """
    return p + p
    pass


df['Name'] = df['Name'].map(aa)
print(df)
# add_data(df, 'AXIS', ws)
# add_data(df, 'HDFC', ws)

wb.save('a.xlsx')
# print(df, df1)

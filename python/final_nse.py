import pandas as pd
import yfinance as yf
from yahoofinancials import YahooFinancials
from nsetools import Nse
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

wb = Workbook()
filename = 'weekly.xlsx'
counter = 0


def add_data(df, symbol):
    global counter

    ws = wb.create_sheet(symbol)
    # ws.append([])
    # ws.append([symbol])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    print(df)
    counter = counter + 1
    if counter % 10 == 0:
        wb.save(filename)
        print('Saved till ', symbol, counter)
# COFORGE.NS: CONFIPET.NS: , DEEPENR.NS: EMAMIPAP.NS:


def fetchData(symbol):
    return yf.download(symbol + '.NS',
                       start=  # '2020-09-08'
                       '1700-01-01',
                       end='2020-12-31',
                       interval="1mo",
                       progress=False)


unableToFetch = []


def map_date(date):

    return date.strftime("%m/%d/%Y")


def get_add(symbol):
    df = fetchData(symbol)

    # print(df.empty, df, df.columns)
    if df.empty:
        unableToFetch.append(symbol)
        return
    df['Date'] = df.index
    df['Date'] = df['Date'].map(map_date)

    date = df['Date']
    df.drop(labels=['Date'], axis=1, inplace=True)
    df.insert(0, 'Date', date)
    # print(df['Open'], df)
    # print(df['Date'], df)
    add_data(df, symbol)


nse = Nse()
fromWhat = 800
fromto = 2000

stocks = list(nse.get_stock_codes())[fromWhat:fromto]
# stocks = ['KIRLOSBROS'
#           # , 'COFORGE', 'CONFIPET', 'DEEPENR', 'EMAMIPAP'
#           ]
print(len(stocks))
# COFORGE.NS: CONFIPET.NS: , DEEPENR.NS: EMAMIPAP.NS:
result = list(map(get_add, stocks))
wb.save(filename)
if len(unableToFetch) != 0:
    ws = wb.create_sheet('UnableToFetch from ' +
                         str(fromWhat) + ' to ' + str(fromto))
    ws.append(unableToFetch)
    wb.save(filename)
    pass
